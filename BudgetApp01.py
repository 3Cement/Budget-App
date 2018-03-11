#! phyton3
# myBudgetApp.py - An app to manage your home budget.
# It is a program that I write while learning programming. I encourage you to comment and write comments about the code.
# Planned versions: web and desktop
import os, xlrd, operator, requests, sys, webbrowser, bs4, time, datetime, re
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.compat import range
from tabulate import tabulate
from operator import itemgetter, attrgetter
#from openpyxl.cell import get_column_letter
main_file = "BudgetFile.xlsx" 
all_transactions = []
categories = {}
def working_directory_data():
	cwd = os.getcwd()
	print(cwd)
	list_of_files = os.listdir('.')
	print(list_of_files)
def yes_or_no(question):
	reply = input(question+' (y/n):\n').lower().strip()
	if reply[:1] == 'y' or reply == '':
		print('Great.')
		return True
	elif reply[:1] == 'n':
		#print('Bad to hear that.')
		return False
	else:
		return yes_or_no('Please type Yes or No ')
def start_program():
	start_again = yes_or_no('Would you like to start again?')
	if start_again == True:
		start()
	else:
		print('We are sorry to hear that.')
def test_excel(user_name):
	filename = create_user_workbook(user_name)
	test_expenditures(filename)
	open_excel_file(filename)
def open_file(filename):
	try:
		print('Oppening main file...')
		os.startfile(filename)
	except Exception as exc:
		print('There was a problem: %s' % (exc))
def id_update(active_user):
	wb = load_workbook(main_file)
	ws = wb[active_user]
	for row_num in range(2, ws.max_row+1):
		ws.cell(row=row_num,column=1).value = row_num-1
	wb.save(main_file)
class User():
	_registry = []
	class_counter = 1
	def __init__(self,name,email,password):
		self._registry.append(self)
		self.id = User.class_counter
		self.name = name
		self.email = email
		self.password = password
		print('User {}.{} created.'.format(self.id,self.name))
		User.class_counter += 1
def budget_file():
	if os.path.isfile(main_file) == True:
		wb = load_workbook(filename=main_file)
		print(main_file+" exist. File opened!")
	else:
		wb = Workbook()
		ws = wb.active
		ws.title = "Users"
		ws.append(["Name", "Email", "Password"])
		ws.freeze_panes = "A2"
		wb.save(filename=main_file)
		print(main_file+" created. File opened!")
		return main_file
def create_username():
	username = str(input('Please add username: '))
	while True:
		try:
			usernameRegex = re.compile(r'''(
				[a-zA-Z0-9._]{3}? 		# username
				)''', re.VERBOSE)
			if not usernameRegex.match(username):
				print('Wrong username format! Please, type your username in correct format.')
				return create_username() # if the user didn't enter right format, try again
			print('Great! Good username format.')
			return username
		except:
			print('Something go wrong!')
			return create_username() # if the user didn't enter right format, try again	
def create_email():
	email = input('Please add user email: ')
	while True:
		try:
			emailRegex = re.compile(r'''(
				[a-zA-Z0-9._%+-]+ 		# username
				@						# @ symbol
				[a-zA-Z0-9.-]+			# domain name
				(\.[a-zA-Z]{2,4})		# dot-something
				)''', re.VERBOSE)
			if not emailRegex.match(email):
				print('Wrong email format! Please, type your email in correct format.')
				return create_email() # if the user didn't enter right format, try again
			print('Great! Good email format.')
			return email
		except:
			return create_email() # if the user didn't enter right format, try again
def create_password():
	password = str(input('Please add password: '))
	while True:
		try:
			usernameRegex = re.compile(r'''(
				[a-zA-Z0-9._]{3}? 		# username
				)''', re.VERBOSE)
			if not usernameRegex.match(password):
				print('Wrong password format! Please, type your password in correct format.')
				return create_password() # if the user didn't enter right format, try again
			print('Great! Good password format.')
			return password
		except:
			print('Something go wrong!')
			return create_password() # if the user didn't enter right format, try again	
'''
def userInputPasswordCheck():
	passwordRegex = re.compile(r'''(
    ^(?=.*[A-Z].*[A-Z])                # at least two capital letters
    (?=.*[!@#$&*])                     # at least one of these special characters
    (?=.*[0-9].*[0-9])                 # at least two numeric digits
    (?=.*[a-z].*[a-z].*[a-z])          # at least three lower case letters
    .{10,}                              # at least 10 total digits
    $
    )''', re.VERBOSE)

    ppass = input("Enter a potential password: ")
    mo = passwordRegex.search(ppass)
    if (not mo):
        print("Password not strong enough")
        return False
    else:
        print("Password long and strong enough")
        return True
'''
def create_user():
	print('Creating new user...')
	#name = input('Please add user name: ')
	#email = input('Please add user email: ')
	name = create_username()
	email = create_email()
	password = create_password()
	new_user = User(name,email,password)
	wb = load_workbook(filename=main_file)
	ws = wb["Users"]
	#ws = wb.active
	ws.append([name,email,password])
	ws1 = wb.create_sheet(title=name)
	ws1.append(["id:", "title", "category", "m_category", "price", "date"])
	ws1.freeze_panes = "A2"
	wb.save(main_file)
	print('User '+name+' created!')
	return new_user
create_user()
def user_names():
	wb = load_workbook(filename=main_file)
	ws = wb["Users"]
	user_names_list = []
	for rowNum in range(2, ws.max_row+1):
		user_name = ws.cell(row=rowNum, column=1).value
		user_names_list.append(user_name)
	print(user_names_list)
	return user_names_list
def user_exist():
	user_names_list = user_names()
	checking_exist = 3
	while checking_exist != 0:
		checking_exist -= 1
		name = input('Please type your account name:\n')
		if name in user_names_list:
			print('Great {}! We found your account!'.format(name))
			return name	
		else:
			print('Wrong username!')
	else:
		print('There is no user with that name. Start again.')
		return False
def show_user_data(name):
	#user_name = input('Please type user name: ')
	wb = load_workbook(filename=main_file)
	worksheet = wb.get_sheet_by_name("Users")
	for rowNum in range(2, worksheet.max_row+1):
		if name == worksheet.cell(row=rowNum, column=1).value:
			print("Great, user "+user_name+" founded!")
			email = worksheet.cell(row=rowNum, column=2).value
			password = worksheet.cell(row=rowNum, column=3).value
			print('User email: '+str(email)+', password: '+str(password))
def user_login(name):
	logging_in = 3
	print('Hello {}! Please log in.'.format(name))
	while logging_in != 0:
		logging_in -= 1
		email = input('Please type your email adress:\n')
		password = input('Please type your password:\n')
		wb = load_workbook(filename=main_file)
		worksheet = wb["Users"]
		for rowNum in range(2, worksheet.max_row+1):
			wb_name = worksheet.cell(row=rowNum, column=1).value
			wb_email = worksheet.cell(row=rowNum, column=2).value
			wb_password = worksheet.cell(row=rowNum, column=3).value
			if wb_name == name and wb_email == email and wb_password == password:
				print('Hello {}! You logged in!'.format(name))
				#break
				ws = wb[name]
				return name
		else:
			print('Wrong email or password!')
	else:
		print('Something go wrong with logging in. Start again')
		return False
def active_user():
	active_user = input("Please put your user name:")
	return active_user
class Transaction():
	""" This is expenditure """
	_registry = []
	class_counter = 1
	def __init__(self,title,category,m_category,price,date):
		self._registry.append(self)
		self.id = Transaction.class_counter
		self.title = title
		self.category = category
		self.m_category = m_category
		self.price = price
		self.date = date
		print('Transaction %s created.' % (title))
		Transaction.class_counter += 1
	def __repr__(self):
		return('\nTransaction nr %s\nTitle: %s\nCategory: %s\nMain category: %s\nPrice: %s$\nDate: %s' % 
			(self.id, self.title, self.category, self.m_category, self.price, self.date))
	def add_to_excel(self,active_user):
		wb = load_workbook(filename=main_file)
		ws = wb[active_user]
		#print('Active user is: '+str(active_user))
		to_excel = [self.id,self.title,self.category,self.m_category,self.price,self.date]
		ws.append(to_excel)
		print(self.title+' added to excel file!')
		wb.save(filename=main_file)
def create_date(active_user):
	reply = yes_or_no("Insert today's date?")
	if reply == True:
		date_now = datetime.datetime.now().strftime('%Y-%m-%d') 
		print('Date now: '+str(date_now))
	else:
		while True:
			try:
				date_entry = input('Enter a date in YYYY-MM-DD format:\n')
				year, month, day = map(int, date_entry.split('-'))
				if year <= 2000 or year >= 2020:
					continue
				date = datetime.date(year,month,day)
				print('Date set to: '+str(date))
				break
			except:
				continue # if the user didn't enter right format, try again
def create_expenditure(active_user):
	title = input('Please add expense title: ')
	category = input('Please add expense category: ')
	m_category = input('Please add expense main category: ')
	price = float(input('Please add price: '))
	date = create_date(active_user)
	new_expenditure = Transaction(title,category,m_category,-price,date)
	new_expenditure.add_to_excel(active_user)
	all_transactions.append(new_expenditure)
def create_expenditure2(active_user,title,category,m_category,price,date):
	new_expenditure = Transaction(title,category,m_category,-price,date)
	new_expenditure.add_to_excel(active_user)
	all_transactions.append(new_expenditure)
def test_expenditures(active_user):
	create_expenditure2(active_user,"Piwko", "Spożywka", "Jedzenie", 2.35, "2016-11-12")
	create_expenditure2(active_user,"Prąd", "Opłaty", "Mieszkanie", 6.5, "2017-12-01")
	create_expenditure2(active_user,"Czynsz", "Meble", "Mieszkanie", 123, "2015-01-25")
	create_expenditure2(active_user,"Bilet", "Autobus", "Transport", 2.5, "2018-01-05")
	create_expenditure2(active_user,"Opłata za taxi", "Taxi", "Transport", 2.5, "2017-02-06")
	create_expenditure2(active_user,"Szynka", "Mięso", "Jedzenie", 2.5, "2018-02-03")
	create_expenditure2(active_user,"Bilet do kina", "Kino", "Rozrywka", 2.5, "2018-03-05")
def create_income(active_user):
	title = input('Please add expense title: ')
	category = input('Please add expense category: ')
	m_category = input('Please add expense main category: ')
	price = float(input('Please add price: '))
	date = create_date(active_user)
	new_income = Transaction(title,category,m_category,price,date)
	new_income.add_to_excel(active_user)
	all_transactions.append(new_income)
def create_income2(active_user,title,category,m_category,price,date):
	new_income = Transaction(title,category,m_category,price,date)
	new_income.add_to_excel(active_user)
	all_transactions.append(new_income)
def test_incomes(active_user):
	create_income2(active_user,"Styczeń", "Wypłata", "Praca", 835, "2018-01-03")
	create_income2(active_user,"Luty", "Wypłata", "Praca", 650, "2018-02-02")
	create_income2(active_user,"Wygrana", "Poker", "Biznes", 18.88, "2018-01-13")
def delete_transaction(active_user,wb):
	reply = input("Select the expenditure to be removed:\n")
	old_sheet = wb[active_user]
	old_sheet.title = active_user+'_old'
	max_row = old_sheet.max_row
	max_col = old_sheet.max_column
	wb.create_sheet(active_user)
	new_sheet = wb[active_user]
	# Do the header.
	for col_num in range(1, max_col+1):
		new_sheet.cell(row=1, column=col_num).value = old_sheet.cell(row=1, column=col_num).value
	for row_num in range(2, max_row):
		if reply == str(row_num):
			print('Transction '+old_sheet.cell(row=(int(reply)+1), column=2).value+' deleted')
			for row_num in range(1, int(reply)+1):
				for col_num in range(1, max_col+1):
					new_sheet.cell(row=row_num,column=col_num).value=old_sheet.cell(row=row_num,column=col_num).value
				#print("Before done!")
			for row_num in range(int(reply), max_row):
				for col_num in range(1, max_col+1):
					new_sheet.cell(row=(row_num+1),column=col_num).value=old_sheet.cell(row=(row_num+2),column=col_num).value
				#print("After done!")
	wb.remove(old_sheet)
	wb.save(main_file)
	id_update(active_user)
def show_transactions(active_user,wb,ws):
	tableData = []
	totalBalance = 0
	print("All transactions:")
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		tableData.append([col[0].value,col[1].value,col[2].value,col[3].value,col[4].value,col[5].value])
		totalBalance += float(col[4].value)
	print(tabulate((tableData),tablefmt="grid"))
	print('Total balance is: $'+str(totalBalance))
def show_transactions_sorted_by(active_user,wb,ws):
	print('Sort transactions by:')
	i = 1
	sortBy = [	"Id",
				"Title",
				"Category",
				"Main Category",
				"Price",
				"Date",
				"Back to main menu"]
	for x in sortBy:
		print(x.ljust(35,'-')+str(i).rjust(2,'-'))
		i += 1
	reply = input('Type number from 1 to {}'.format(i-1).center(37,'-')+'\n')
	if reply == "7":
		what_next(active_user)
		print("Returning to the main menu...")
	else:
		tableData = []
		for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
			tableData.append((col[0].value,col[1].value,col[2].value,col[3].value,col[4].value,col[5].value))
		sortedData = sorted(tableData, key=operator.itemgetter(int(reply)-1))
		header = ("id","Title","Category","m_Category","Price","Date")
		sortedData.insert(0,header)
		print('Transactions sorted by: '+header[int(reply)-1])
		print(tabulate((sortedData),tablefmt="grid"))
def show_incomes(active_user,wb,ws):
	tableData = []
	allIncomes = 0
	print("All incomes:")
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if float(col[4].value) > 0:
			tableData.append([col[0].value,col[1].value,col[2].value,col[3].value,col[4].value,col[5].value])
			allIncomes += float(col[4].value)
	header = ("id","Title","Category","m_Category","Price","Date")
	tableData.insert(0,header)
	print(tabulate((tableData),tablefmt="grid"))
	print('The sum of incomes is: $'+str(allIncomes))
def show_expenditures(active_user,wb,ws):
	tableData = []
	allExpenditures = 0
	print("All expenditures:")
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if float(col[4].value) < 0:
			tableData.append([col[0].value,col[1].value,col[2].value,col[3].value,col[4].value,col[5].value])
			allExpenditures += float(col[4].value)
	header = ("id","Title","Category","m_Category","Price","Date")
	tableData.insert(0,header)
	print(tabulate((tableData),tablefmt="grid"))
	print('The sum of expenditures is: $'+str(allExpenditures))
def show_categories(active_user,wb,ws):
	categories = []
	print("All expenditures:")
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if str(col[2].value) not in categories:
			categories.append(col[2].value)
	print('List of categories: '+str(categories))
def show_main_categories(active_user,wb,ws):
	mainCategories = {}
	print("All expenditures:")
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if str(col[3].value) not in mainCategories.keys():
			key = str(col[3].value)
			mainCategories.setdefault(key, [])
			mainCategories[key].append(str(col[2].value))
	for col in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
		if str(col[2].value) not in mainCategories[str(col[3].value)]:
			mainCategories[str(col[3].value)].append(str(col[2].value))
	print('List of main categories with subcategories:\n'+str(mainCategories))
def income_or_expense(question,active_user):
	selection = input(question+' (I or E):\n').lower().strip()
	if selection[:1] == 'i' or selection == '':
		create_income(active_user)
	elif selection[:1] == 'e':
		create_expenditure(active_user)
	else:
		return income_or_expense('Please type I (for income) or E (for expenditure)')
def start():
	#working_directory_data()
	budget_file()
	print("Existing users: ")
	user_names()
	reply = yes_or_no('Hello. Do you already have user account?')
	if reply == True:
		name = user_exist()
		if name == False:
			start_program()
		else:
			user_login(name)
			print(name)
			test_expenditures(name)
			test_incomes(name)
			id_update(name)
			what_next(name)
			#open_file(main_file)
	elif reply == False:
		print('Then you need to create new user: ')
		name = user_login(create_user().name)
		print(name)
		test_expenditures(name)
		test_incomes(name)
		what_next(name)
		#open_file(main_file)
	else:
		print('Something go wrong with your name!')
def what_next(active_user):
	wb = load_workbook(main_file)
	ws = wb[active_user]
	print('Please choose what would you like to do next')
	i = 1
	strings = [	"Show all transcations",
				"Add new transaction",
				"Delete transaction",
				"Show only expenditures",
				"Show only incomes",
				"Show transactions in correct order",
				"Show transactions by date",
				"Show main categories",
				"Show categories",
				"Show charts",
				"Restart program",
				"Open budget file",
				"Save and exit"]
	for string in strings:
		print(string.ljust(35,'-')+str(i).rjust(2,'-'))
		i += 1
	reply = input('Type number from 1 to {}'.format(i-1).center(37,'-')+'\n')
	if reply == "1":
		show_transactions(active_user,wb,ws)
		what_next(active_user)
	elif reply == "2":
		income_or_expense("Would you like to add income or expanditure?",active_user)
		what_next(active_user)
	elif reply == "3":
		print("Delete transaction:")
		show_transactions(active_user,wb,ws)
		delete_transaction(active_user,wb)
		what_next(active_user)
	elif reply == "4":
		show_expenditures(active_user,wb,ws)
		what_next(active_user)
	elif reply == "5":
		show_incomes(active_user,wb,ws)
		what_next(active_user)
	elif reply == "6":
		print("Showing transaction. Choose order:")
		show_transactions_sorted_by(active_user,wb,ws)
		what_next(active_user)
	elif reply == "7":
		print("Showing transactions by date:")
	elif reply == "8":
		print("Showing main categories:")
		show_main_categories(active_user,wb,ws)
		what_next(active_user)
	elif reply == "9":
		print("Showing categories:")
		show_categories(active_user,wb,ws)
		what_next(active_user)
	elif reply == "10":
		print("Show charts. Choose data:")
	elif reply == "11":
		print("Restarting program:")
		start()
	elif reply == "12":
		open_file(main_file)
	elif reply == "13":
		print("Saving data and exit...")
		wb.save(main_file)
		exit()
	else:
		print("Error. Something go wrong...")
	open_file(main_file)

start()
